import argparse
import json
import logging
from pathlib import Path
from urllib.parse import quote_plus

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

ROOT = Path(__file__).resolve().parents[2]
VALIDATION_DIR = ROOT / "validation"
DEFAULT_OUT_DIR = VALIDATION_DIR / "blinded_agent_package"
GT_BATCH_INPUT = ROOT / "data" / "batches" / "ready_for_classification" / "1000batch_api.jsonl"
NOTE_TAGS = "`paper-reviewed`, `abstract-only`, `domain-knowledge`, or `unresolved-after-review`"

DATASET_ALLOWED = [
    (
        "annotator_Modalities",
        "Text; Audio; Image; Video; Time series; Graph; Spatial; Multimodal",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_TaskTypes",
        "Classification; Regression; Sequence labeling; Generation; Summarization; Translation; Question answering; Reasoning; Dialogue; Object detection; Forecasting; Retrieval; Alignment; Multimodal integration; Clustering; Reinforcement learning",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_Domains",
        "General; Media; Scientific / academic; Healthcare; Legal; Economics; Social; Geospatial; Robotics; Vision; Entertainment; Education; Infrastructure; Ontology; Biology; Chemistry; Environmental",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_Annotation",
        "Fully Supervised; Weakly Supervised; Self-Supervised; Semi-Supervised; Reinforcement Feedback; Crowdsourced; Expert Annotations",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_Size",
        "Small; Medium; Large",
        "Single label. Small <10K items, Medium 10K-100K, Large >100K.",
    ),
    (
        "annotator_Granularity",
        "Document-level; Sentence-level; Token-level; Frame-level; Pixel-level; Object-level",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_Linguistic",
        "Monolingual: English, Chinese, Spanish, French, German, Russian, Portuguese, Italian, Dutch, Arabic, Japanese, Korean, Turkish, Polish, Vietnamese, Indonesian, Hebrew, Swedish, Czech, Hungarian, Other; Multilingual: <languages>; Cross-lingual: <language pair>",
        "Use the exact format `Monolingual: English`, `Multilingual: English, French`, or `Cross-lingual: English ↔ French`.",
    ),
    (
        "annotator_CognitiveAffective",
        "Attention; Memory; Problem Solving; Reasoning; Decision Making; Perception; Learning; Cognitive Load; Emotion; Empathy; Theory of Mind; Social Reasoning; Moral Cognition; Personality",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_DataQuality",
        "Noisy; Curated",
        "Multi-label allowed if both clearly apply; otherwise choose the dominant label and explain in notes.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with one evidence tag: {NOTE_TAGS}, then give a short justification.",
    ),
]

MODEL_ALLOWED = [
    (
        "annotator_Architecture",
        "Transformer (Encoder); Transformer (Decoder); Transformer (Encoder-Decoder); Generative; Convolutional (CNN); Recurrent (RNN / LSTM); Graph Neural Network (GNN); Tree-based; Linear; Kernel Models; Probabilistic; Reinforcement Learning (RL)",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_TrainingParadigm",
        "Supervised Learning; Self-supervised Learning; Unsupervised Learning; Reinforcement Learning; Multi-task Learning; Few-shot Learning; Zero-shot Learning; Fine-tuning; Retrieval-Augmented Generation (RAG)",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_Provider",
        "OpenAI; Anthropic; Meta AI; Google DeepMind; Mistral AI; Alibaba / Qwen team; Cohere; Hugging Face; Stability AI; Microsoft Research; NVIDIA / NeMo; Databricks / MosaicML; DeepSeek; Other / Academic",
        "Single label.",
    ),
    (
        "annotator_Openness",
        "Closed; Open",
        "Single label.",
    ),
    (
        "annotator_Size",
        "Small (<1B); Medium (1-10B); Large (10-100B); Extra-Large (>100B)",
        "Single label. Fill the best-supported size after review. Leave blank only if the paper was checked and size is genuinely not inferable.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with one evidence tag: {NOTE_TAGS}, then give a short justification.",
    ),
]

METRIC_ALLOWED = [
    (
        "annotator_EvType",
        "Accuracy; Ranking; Regression; Continuous Prediction; Probability; Uncertainty; Fairness; Safety; Efficiency / Latency; Explainability; Robustness; User Experience",
        "Multi-label allowed. Separate multiple labels with semicolons.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with one evidence tag: {NOTE_TAGS}, then give a short justification.",
    ),
]

INTRODUCEDNESS_ALLOWED = [
    (
        "annotator_introducedness",
        "pre-existing_reusable; paper-introduced; paper-specific_derivative; unclear",
        "Use `unclear` only after reviewing the paper when the status is still genuinely unresolved.",
    ),
    (
        "annotator_mismatch_type",
        "alias_variant; same_family; paper_specific; established_but_absent; unclear",
        "If `annotator_introducedness` is `paper-introduced` or `paper-specific_derivative`, default to `paper_specific`. If it is `pre-existing_reusable` and there is no clear alias or family evidence, default to `established_but_absent`.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with one evidence tag: {NOTE_TAGS}, then give a short justification.",
    ),
]

EXTRACTION_ALLOWED = [
    (
        "annotator_num_correct",
        "Integer from 0 to num_extracted",
        "Count extracted entities genuinely used in the paper's experiments, evaluations, or ablations.",
    ),
    (
        "annotator_num_hallucinated",
        "Integer from 0 to num_extracted",
        "Count extracted entities that do not appear in the paper's experiments.",
    ),
    (
        "annotator_missed_entities",
        "Semicolon-separated entity names",
        "List experimentally used entities missing from `extracted_entities`.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with one evidence tag: {NOTE_TAGS}, then give a short justification.",
    ),
]

NORMALIZATION_ALLOWED = [
    (
        "annotator_should_merge",
        "yes; no",
        "Answer `yes` only if both names refer to the same underlying entity rather than different versions, sizes, or fine-tunes.",
    ),
    (
        "annotator_notes",
        "Free text",
        f"Required for every row. Start with `domain-knowledge` or `unresolved-after-review`, then give a short justification.",
    ),
]


def parse_titles(path: Path) -> dict[str, str]:
    titles: dict[str, str] = {}
    with path.open(encoding="utf-8") as handle:
        for line in handle:
            obj = json.loads(line)
            custom_id = obj["custom_id"]
            input_text = obj["body"].get("input", "")
            marker = "Title: "
            start = input_text.find(marker)
            if start < 0:
                continue
            rest = input_text[start + len(marker):]
            end = rest.find("\n")
            titles[custom_id] = rest[:end].strip() if end >= 0 else rest.strip()
    return titles


def arxiv_search_url(title: str) -> str:
    if not title:
        return ""
    return f"https://arxiv.org/search/?query={quote_plus(title)}&searchtype=all"


def add_record_id(df: pd.DataFrame, prefix: str) -> pd.DataFrame:
    out = df.copy()
    out.insert(0, "record_id", [f"{prefix}_{i:04d}" for i in range(1, len(out) + 1)])
    return out


def enrich_titles(df: pd.DataFrame, titles: dict[str, str]) -> pd.DataFrame:
    out = df.copy()
    if "paper_id" in out.columns:
        out.insert(1 if "record_id" in out.columns else 0, "paper_title", out["paper_id"].map(titles).fillna(""))
        out.insert(2 if "record_id" in out.columns else 1, "arxiv_search_url", out["paper_title"].map(arxiv_search_url))
    return out


def to_text_df(path: Path) -> pd.DataFrame:
    return pd.read_csv(path).fillna("")


def build_extraction_df(titles: dict[str, str]) -> pd.DataFrame:
    df = to_text_df(VALIDATION_DIR / "validation_extraction.csv").rename(columns={"title": "paper_title"})
    df["arxiv_search_url"] = df["paper_title"].map(arxiv_search_url)
    cols = [
        "paper_id",
        "paper_title",
        "arxiv_search_url",
        "research_question",
        "entity_type",
        "extracted_entities",
        "num_extracted",
        "annotator_num_correct",
        "annotator_num_hallucinated",
        "annotator_missed_entities",
        "annotator_notes",
    ]
    return add_record_id(df[cols], "extraction")


def build_classification_df(entity_kind: str, titles: dict[str, str]) -> pd.DataFrame:
    df = to_text_df(VALIDATION_DIR / f"validation_classification_{entity_kind}.csv")
    keep = ["paper_id", "entity_name"]
    keep.extend([c for c in df.columns if c.startswith("annotator_")])
    blinded = df[keep]
    blinded = add_record_id(blinded, f"classification_{entity_kind}")
    blinded = enrich_titles(blinded, titles)
    front = ["record_id", "paper_id", "paper_title", "arxiv_search_url", "entity_name"]
    rest = [c for c in blinded.columns if c not in front]
    return blinded[front + rest]


def build_normalization_df() -> pd.DataFrame:
    df = to_text_df(VALIDATION_DIR / "validation_normalization.csv")
    keep = ["entity_type", "representative", "variant", "annotator_should_merge", "annotator_notes"]
    return add_record_id(df[keep], "normalization")


def build_introducedness_df(titles: dict[str, str]) -> pd.DataFrame:
    df = to_text_df(VALIDATION_DIR / "validation_introducedness_R1.csv")
    keep = ["paper_id", "entity_type", "entity_name", "annotator_introducedness", "annotator_mismatch_type", "annotator_notes"]
    blinded = add_record_id(df[keep], "introducedness")
    blinded = enrich_titles(blinded, titles)
    front = ["record_id", "paper_id", "paper_title", "arxiv_search_url", "entity_type", "entity_name"]
    rest = [c for c in blinded.columns if c not in front]
    return blinded[front + rest]


def instructions_df(task: str) -> pd.DataFrame:
    rows: list[tuple[str, str]]
    if task == "extraction":
        rows = [
            ("Scope", "Use only this workbook, the package README, and the linked arXiv paper. This workbook is one of six; complete all six overall, but work on one workbook at a time."),
            ("What is shown", "`extracted_entities` is the output being audited. No reference answers are shown."),
            ("Paper use", "Group rows by `paper_id`. Fetch each paper once, then annotate all rows for that paper from the paper text or abstract."),
            ("How to fill", "Fill only `annotator_num_correct`, `annotator_num_hallucinated`, `annotator_missed_entities`, and `annotator_notes`."),
            ("Completion rule", "Complete every row. Do not leave extraction counts blank."),
            ("Counting rule", "`annotator_num_correct + annotator_num_hallucinated` should usually equal `num_extracted`."),
            ("Missed entities", "Write missed entities as a semicolon-separated list."),
            ("Notes rule", f"Every row needs `annotator_notes` starting with one evidence tag: {NOTE_TAGS}."),
        ]
    elif task.startswith("classification_"):
        rows = [
            ("Scope", "Use only this workbook, the package README, and the linked arXiv paper. This workbook is one of six; complete all six overall, but work on one workbook at a time."),
            ("What is shown", "The workbook shows the paper context and entity name, but hides all pipeline labels and source metadata."),
            ("Paper use", "Group rows by `paper_id`. Fetch each paper once, then annotate all rows for that paper together. Normalization is the only task that is normally name-only."),
            ("How to fill", "Fill only the `annotator_*` columns and `annotator_notes`."),
            ("Completion rule", "After reviewing the paper, fill the best-supported label for each field. Leave a field blank only if the paper was checked and the field is genuinely not inferable."),
            ("Multi-label rule", "When multiple labels apply, separate them with semicolons and use exact spellings from the `allowed_values` sheet."),
            ("Single-label rule", "For single-label fields, enter one value only. If the field remains unresolved after paper review, leave it blank and explain that in notes with `unresolved-after-review`."),
            ("Evidence rule", "Use `paper-reviewed` by default. Use `abstract-only` only if HTML is unavailable. Use `domain-knowledge` only for well-known, unambiguous entities."),
            ("Notes rule", f"Every row needs `annotator_notes` starting with one evidence tag: {NOTE_TAGS}."),
        ]
    elif task == "normalization":
        rows = [
            ("Scope", "Use only this workbook and the package README. This workbook is one of six; complete all six overall, but work on one workbook at a time. Do not inspect the original normalization CSV."),
            ("What is shown", "The workbook hides the pipeline merge decision and fuzzy score."),
            ("How to fill", "Fill only `annotator_should_merge` with `yes` or `no`, plus `annotator_notes` on every row."),
            ("Completion rule", "Answer every row with `yes` or `no`. Do not leave merge decisions blank."),
            ("Decision rule", "Aliases and trivial spelling variants should merge; different versions, sizes, modalities, or fine-tunes should not."),
            ("Notes rule", "Every row needs `annotator_notes` starting with `domain-knowledge` or `unresolved-after-review`, then a short justification."),
        ]
    else:
        rows = [
            ("Scope", "Use only this workbook, the package README, and the linked arXiv paper. This workbook is one of six; complete all six overall, but work on one workbook at a time."),
            ("What is shown", "The workbook hides all heuristic hints such as singleton flags, frequency counts, title matches, and provider cues."),
            ("Paper use", "Group rows by `paper_id`. Fetch each paper once, then annotate all rows for that paper together."),
            ("How to fill", "Fill only `annotator_introducedness`, `annotator_mismatch_type`, and `annotator_notes`."),
            ("Completion rule", "Complete every row. Use `unclear` only after reviewing the paper when the status is still genuinely unresolved."),
            ("Mismatch rule", "If `annotator_introducedness` is `paper-introduced` or `paper-specific_derivative`, default `annotator_mismatch_type` to `paper_specific` unless there is strong evidence otherwise."),
            ("Pre-existing rule", "If `annotator_introducedness` is `pre-existing_reusable` and there is no clear alias or family evidence, default `annotator_mismatch_type` to `established_but_absent` rather than `unclear`."),
            ("Notes rule", f"Every row needs `annotator_notes` starting with one evidence tag: {NOTE_TAGS}."),
        ]
    return pd.DataFrame(rows, columns=["rule", "details"])


def allowed_values_df(task: str) -> pd.DataFrame:
    if task == "extraction":
        rows = EXTRACTION_ALLOWED
    elif task == "classification_datasets":
        rows = DATASET_ALLOWED
    elif task == "classification_models":
        rows = MODEL_ALLOWED
    elif task == "classification_metrics":
        rows = METRIC_ALLOWED
    elif task == "normalization":
        rows = NORMALIZATION_ALLOWED
    else:
        rows = INTRODUCEDNESS_ALLOWED
    return pd.DataFrame(rows, columns=["field", "allowed_values", "notes"])


def style_workbook(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for ws in wb.worksheets:
        if ws.max_row == 0 or ws.max_column == 0:
            continue
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(vertical="top", wrap_text=True)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)
        headers = [cell.value for cell in ws[1]]
        if "arxiv_search_url" in headers:
            col_idx = headers.index("arxiv_search_url") + 1
            for row_idx in range(2, ws.max_row + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    cell.hyperlink = str(cell.value)
                    cell.style = "Hyperlink"
        for idx, column_cells in enumerate(ws.columns, start=1):
            max_len = max(len(str(cell.value)) if cell.value is not None else 0 for cell in column_cells)
            width = min(max(max_len + 2, 12), 60)
            ws.column_dimensions[get_column_letter(idx)].width = width
    wb.save(path)


def write_workbook(path: Path, annotate: pd.DataFrame, task: str) -> None:
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        annotate.to_excel(writer, sheet_name="annotate", index=False)
        instructions_df(task).to_excel(writer, sheet_name="instructions", index=False)
        allowed_values_df(task).to_excel(writer, sheet_name="allowed_values", index=False)
    style_workbook(path)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUT_DIR)
    args = parser.parse_args()

    titles = parse_titles(GT_BATCH_INPUT)
    out_dir = args.output_dir
    out_dir.mkdir(parents=True, exist_ok=True)

    workbooks = {
        "extraction_blinded.xlsx": (build_extraction_df(titles), "extraction"),
        "classification_datasets_blinded.xlsx": (build_classification_df("datasets", titles), "classification_datasets"),
        "classification_models_blinded.xlsx": (build_classification_df("models", titles), "classification_models"),
        "classification_metrics_blinded.xlsx": (build_classification_df("metrics", titles), "classification_metrics"),
        "normalization_blinded.xlsx": (build_normalization_df(), "normalization"),
        "introducedness_blinded.xlsx": (build_introducedness_df(titles), "introducedness"),
    }

    for filename, (df, task) in workbooks.items():
        path = out_dir / filename
        write_workbook(path, df, task)
        log.info("Wrote %s", path)


if __name__ == "__main__":
    main()
